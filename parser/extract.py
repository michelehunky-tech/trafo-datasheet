"""Deterministic Excel parser for the Trafo Elettro datasheet portal.

Reads the fixed key-value extraction sheet by LABEL (not by fixed coordinate),
derives the transformer family, applies the canonical schema, translates
enumerated text values, formats numbers, and selects the configuration image.

No LLM is involved. All logic is deterministic.
"""
from pathlib import Path
import yaml
from openpyxl import load_workbook

from .format import format_value, is_blank, looks_like_coerced_date

SCHEMA_PATH = Path(__file__).with_name("schema.yaml")

SECTION_ORDER = ["general", "electrical", "ratings", "environmental", "dimensions"]
SECTION_TITLE = {
    "general": "General",
    "electrical": "Electrical",
    "ratings": "Ratings",
    "environmental": "Working Conditions",
    "dimensions": "Dimensions & weight",
}


def load_schema(path=SCHEMA_PATH):
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


I18N_DIR = Path(__file__).resolve().parent.parent / "i18n"
_LANG_CACHE = {}


def load_lang(code="en"):
    code = (code or "en").lower()
    if code not in _LANG_CACHE:
        p = I18N_DIR / f"{code}.yaml"
        if not p.exists():
            p = I18N_DIR / "en.yaml"
        with open(p, encoding="utf-8") as f:
            _LANG_CACHE[code] = yaml.safe_load(f)
    return _LANG_CACHE[code]


def load_schema_lang(lang_code="en", path=SCHEMA_PATH):
    """Schema con glossario lingua iniettato in schema['lang']."""
    schema = load_schema(path)
    schema["lang"] = load_lang(lang_code)
    return schema


def _nf(schema):
    """Number format della lingua attiva, con fallback a quello dello schema."""
    lg = schema.get("lang")
    if lg and lg.get("number_format"):
        return lg["number_format"]
    return schema["meta"]["number_format"]


def _label(schema, key, fallback=None):
    """Traduce un'etichetta campo (chiave IT dallo schema o EN per righe tabella)."""
    lg = schema.get("lang")
    if lg:
        return lg["fields"].get(key, fallback if fallback is not None else key)
    return fallback if fallback is not None else key


def _section_title(schema, key, fallback):
    lg = schema.get("lang")
    if lg:
        return lg["sections"].get(key, fallback)
    return fallback


def read_raw(xlsx_path, schema):
    """Return (raw, accessories): key-value fields read by label, and the
    accessory list (rows after the 'Accessori:' marker)."""
    meta = schema["meta"]
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[meta["sheet"]] if meta["sheet"] in wb.sheetnames else wb[wb.sheetnames[0]]
    marker = meta.get("accessories_marker", "Accessori:")
    raw, accessories, in_acc = {}, [], False
    for r in range(meta["first_row"], ws.max_row + 1):
        label = ws.cell(r, meta["label_col"]).value
        value = ws.cell(r, meta["value_col"]).value
        if label is None or not str(label).strip():
            continue
        label = str(label).strip()
        if label == marker:
            in_acc = True
            continue
        if in_acc:
            accessories.append(label)          # accessory descriptions are full sentences
        else:
            raw[label] = value
    return raw, accessories


def derive_family(raw, schema):
    """oil if 'Tipologia olio' present; resin if air cooling; else oil."""
    rules = schema["family_rules"]
    if not is_blank(raw.get("Tipologia olio")):
        return "olio"
    cooling = str(raw.get("Raffreddamento") or "").strip().upper()
    if cooling in [c.upper() for c in rules["air_cooling"]]:
        return "resina"
    if cooling[:1] in rules["oil_prefix"]:
        return "olio"
    serie = str(raw.get("Serie") or "")
    return "resina" if "-R" in serie.upper() else "olio"


IMG_DIR = Path(__file__).resolve().parent.parent / "assets" / "transformers"


def _img_exists(key):
    return bool(key) and (IMG_DIR / f"{key}.png").exists()


def _winding_pair(windings):
    """Coppia di sigle senza numeri (es. 'MV-LV', 'MV-MV', 'HV-MV'); con 3 avvolgimenti
    prende primario + primo secondario."""
    if not windings:
        return "MV-MV"
    if len(windings) == 1:
        b = windings[0]["base"]
        return f"{b}-{b}"
    return f"{windings[0]['base']}-{windings[1]['base']}"


def select_image(raw, family, schema):
    """Assegna l'immagine trasformatore secondo la regola concordata.
    Priorità: resina → earthing → olio doppio secondario → olio singolo secondario.
    Ricade sui fallback previsti se la combinazione esatta non esiste."""
    windings = build_windings(raw, schema)
    secondaries = windings[1:]                    # tutti gli avvolgimenti dopo il primario
    n_sec = len(secondaries)
    has_bt = any(w["base"] == "LV" for w in secondaries)
    earthing = is_earthing(raw)
    casa = str(raw.get("Tipo casa") or "").strip().lower()
    cooling = str(raw.get("Tipo sistema raffreddamento") or "").strip().lower()
    oltc = "sottocarico" in str(raw.get("Tipo commutatore") or "").lower()
    ermetico = "ermetico" in casa
    radiatori = "radiatori" in cooling
    scambiatore = "scambiatore" in cooling
    p_mt = _f(raw.get("Potenza nominale MT"))

    cand = []

    # 1) RESINA (ma earthing ha priorità: gestito sotto)
    if earthing:
        if family == "resina":
            cand.append("Earthing_resina")
        else:
            cassa = "ermetico" if ermetico else "conservatore"
            if has_bt:
                cand.append(f"Earthing_{cassa}_avvolgimento_BT")
            cand.append(f"Earthing_{cassa}")
        cand.append("Earthing_resina")

    elif family == "resina":
        cand.append("Trasformatore_resina")

    # 2) OLIO, DOPPIO SECONDARIO (>=2 secondari): solo cassa+raffreddamento
    elif n_sec >= 2:
        if ermetico:
            cand.append("Trasformatore_in_olio_ermetico_DOPPIO_SECONDARIO")
        elif scambiatore:
            cand.append("Trasformatore_in_olio_conservatore_e_scambiatore_OFWF_DOPPIO_SECONDARIO")
        elif radiatori:
            cand.append("Trasformatore_in_olio_conservatore_radiatori_DOPPIO_SECONDARIO")
        else:  # conservatore onde / generico
            cand.append("Trasformatore_in_olio_conservatore_DOPPIO_SECONDARIO")
        cand.append("Trasformatore_in_olio_conservatore_DOPPIO_SECONDARIO")

    # 3) OLIO, SINGOLO SECONDARIO
    else:
        if scambiatore:
            cand.append("Trasformatore_in_olio_conservatore_e_scambiatore_OFWF")
        elif radiatori and not ermetico:
            sig = _winding_pair(windings)
            comm = "OLTC" if oltc else "a_vuoto"
            if sig == "MV-LV" and not oltc and p_mt is not None and p_mt < 3150:
                cand.append("Trasformatore_in_olio_Conservatore_radiatori_MV-LV_"
                            "commutatore_a_vuoto_per_potenze_ridotte_Potenza_minore_di_3150_kVA")
            elif sig in ("HV-MV", "MV-MV"):
                cand.append(f"Trasformatore_in_olio_Conservatore_radiatori_{sig}_commutatore_{comm}")
            # fallback radiatori: MV-MV a vuoto
            cand.append("Trasformatore_in_olio_Conservatore_radiatori_MV-MV_commutatore_a_vuoto")
        elif ermetico:
            cand.append("Ermetico_onde")
        else:  # conservatore onde
            cand.append("Trasformatore_in_olio_Conservatore_onde")
        cand.append("Trasformatore_in_olio_Conservatore_onde")

    for c in cand:
        if _img_exists(c):
            return c
    return cand[0] if cand else None


def translate(label, raw_value, schema):
    """Traduce un valore enumerato nella lingua attiva (schema['lang']['values']),
    con fallback ai valori EN dello schema. translated_ok=False = da confermare nel form."""
    vmap = schema.get("value_map", {})
    lookup = label
    if label.startswith("Materiale "):
        lookup = "Materiale MT"
    elif label.startswith("Tipo avvolg. "):
        lookup = "Tipo avvolg. MT"
    if lookup not in vmap or is_blank(raw_value):
        return None, True  # campo non enumerato
    key = str(raw_value).strip()
    lg = schema.get("lang")
    src = lg["values"] if lg else vmap[lookup]
    if key in src:
        return src[key], True
    toks = key.split()
    if len(toks) > 1 and toks[0] in src:
        return f"{src[toks[0]]} {' '.join(toks[1:])}", True
    return key, False  # valore presente ma non mappato -> conferma nel form


def build_fields(raw, family, schema):
    """Produce the ordered, sectioned list of display fields for the family."""
    nf = _nf(schema)
    out = []
    for fld in schema["fields"]:
        if not fld.get("include_in_sheet"):
            continue
        if fld["family"] not in ("both", family):
            continue
        if fld["section"] == "header":
            continue
        raw_v = raw.get(fld["it"])
        # text translation (only for mapped fields)
        translated, ok = translate(fld["it"], raw_v, schema)
        display = translated if translated is not None else format_value(raw_v, fld, nf)
        if fld["it"] == "Colore" and not is_blank(raw_v):
            lg = schema.get("lang")
            if lg and lg.get("color_terms"):
                display = translate_color(raw_v, lg["color_terms"])
        blank = is_blank(raw_v)
        # Application forzata a "Earthing transformer" se trasformatore di terra
        if fld["it"] == "Applicazione" and is_earthing(raw):
            lg = schema.get("lang")
            et = (lg["values"] if lg else {}).get("Trasformatore di terra", "Earthing transformer")
            display, ok, blank = et, True, False
        # Load losses = 0 -> ometti il campo
        if fld["it"] in ("Perdite a carico 75°C", "Perdite a carico 120°C"):
            if _f(raw_v) in (0, None):
                blank = True
        row = {
            "it": fld["it"], "en": _label(schema, fld["it"], fld["en"]),
            "section": fld["section"],
            "unit": fld.get("unit"), "value": display,
            "blank": blank, "translated_ok": ok,
        }
        if fld["it"] == "LpA":
            row["newrow"] = True   # LpA/LwA su riga nuova
        out.append(row)
    return out


def grouped_sections(fields, schema=None):
    """Group included, non-blank fields into ordered sections for the template."""
    groups = []
    n = 0
    for sec in SECTION_ORDER:
        rows = [f for f in fields if f["section"] == sec and not f["blank"] and f["value"] not in (None, "")]
        if rows:
            n += 1
            title = _section_title(schema, sec, SECTION_TITLE[sec]) if schema else SECTION_TITLE[sec]
            groups.append({"key": sec, "title": title, "rows": rows, "idx": n})
    return groups


# --- multi-winding configuration ---
WINDINGS = [
    ("MT",  {"V": "Tensione MT", "V2": "Tensione MT2", "P": "Potenza nominale MT",
             "Pf": "Potenza reg. forzato MT",
             "conn": "Collegamento MT", "ins": "Classe isolamento MT", "mat": "Materiale MT",
             "wt": "Tipo avvolg. MT", "tc": "Classe termica MT", "tr": "Sovratemperatura avvolg. MT"}),
    ("BT1", {"V": "Tensione BT1", "P": "Potenza nominale BT1",
             "Pf": "Potenza reg. forzato BT1",
             "conn": "Collegamento BT1", "ins": "Classe isolamento BT1", "mat": "Materiale BT1",
             "wt": "Tipo avvolg. BT1", "tc": "Classe termica BT1", "tr": "Sovratemperatura avvolg. BT1"}),
    ("BT2", {"V": "Tensione BT2", "P": "Potenza nominale BT2",
             "Pf": "Potenza reg. forzato BT2",
             "conn": "Collegamento BT2", "ins": "Classe isolamento BT2", "mat": "Materiale BT2",
             "wt": "Tipo avvolg. BT2", "tc": "Classe termica BT2", "tr": "Sovratemperatura avvolg. BT2"}),
    ("BT3", {"V": "Tensione BT3", "P": "Potenza nominale BT3",
             "Pf": "Potenza reg. forzato BT3",
             "conn": "Collegamento BT3", "ins": "Classe isolamento BT3", "mat": "Materiale BT3",
             "wt": "Tipo avvolg. BT3", "tc": "Classe termica BT3", "tr": "Sovratemperatura avvolg. BT3"}),
]
WINDING_ROWS = [
    ("Rated power", "P", "kVA", 0),
    ("Voltage", "V", "V", 0),
    ("Connection", "conn", None, None),
    ("Insulation level", "ins", "kV", None),
    ("Winding material", "mat", None, None),
    ("Winding type", "wt", None, None),
    ("Thermal class", "tc", None, None),
    ("Winding temp. rise", "tr", "°C", 0),
]
IMPEDANCES = [
    ("Impedenza di cortocircuito % MT-BT1", "MT", "BT1"),
    ("Impedenza di cortocircuito % MT-BT2", "MT", "BT2"),
    ("Impedenza di cortocircuito % MT-BT3", "MT", "BT3"),
    ("Impedenza di cortocircuito % BT1-BT2", "BT1", "BT2"),
    ("Impedenza di cortocircuito % BT1-BT3", "BT1", "BT3"),
    ("Impedenza di cortocircuito % BT2-BT3", "BT2", "BT3"),
]


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _imp(v, nf):
    if is_blank(v):
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return format_value(v, {"decimals": 2}, nf)
    return str(v).strip()  # valori testuali: NA, < 3


def build_windings(raw, schema):
    """Avvolgimenti presenti con sigla LV/MV/HV (numerata se la classe si ripete)."""
    vc = schema["voltage_class"]
    present = []
    for role, m in WINDINGS:
        v = _f(raw.get(m["V"]))
        if v is None or v == 0:
            continue
        ref = v
        v2 = _f(raw.get(m.get("V2")))
        if v2:
            ref = max(v, v2)  # doppia tensione: classifica sulla massima
        base = "LV" if ref <= vc["lv_max"] else ("MV" if ref <= vc["mv_max"] else "HV")
        present.append({"role": role, "base": base, "m": m})
    from collections import Counter
    cnt = Counter(w["base"] for w in present)
    seen = {}
    for w in present:
        b = w["base"]
        if cnt[b] > 1:
            seen[b] = seen.get(b, 0) + 1
            w["label"] = f"{b}{seen[b]}"
        else:
            w["label"] = b
    return present


def _field_by_it(schema, it):
    return next((f for f in schema["fields"] if f["it"] == it), {"it": it})


def get_display(it_label, raw, schema):
    """Single display value for an italian label (translate then format)."""
    fld = _field_by_it(schema, it_label)
    translated, _ = translate(it_label, raw.get(it_label), schema)
    if translated is not None:
        return translated
    return format_value(raw.get(it_label), fld, _nf(schema))


def build_ratings(raw, schema):
    nf = _nf(schema)
    windings = build_windings(raw, schema)
    labels = [w["label"] for w in windings]
    rolelabel = {w["role"]: w["label"] for w in windings}

    rows = []
    for en, key, unit, dec in WINDING_ROWS:
        cells = []
        for w in windings:
            m = w["m"]
            raw_v = raw.get(m[key])
            if dec is not None and isinstance(raw_v, (int, float)) and not isinstance(raw_v, bool):
                val = format_value(raw_v, {"decimals": dec}, nf)
            else:
                val = get_display(m[key], raw, schema)
            if key == "P":
                pf = raw.get(m.get("Pf"))
                pfv, pv = _f(pf), _f(raw_v)
                if pfv not in (None, 0) and pfv != pv:
                    val = (f"{format_value(raw_v, {'decimals':0}, nf)}/"
                           f"{format_value(pf, {'decimals':0}, nf)}")
            if key == "V" and w["role"] == "MT":            # doppia tensione MT: min / max
                v1, v2 = _f(raw.get(m["V"])), _f(raw.get(m.get("V2")))
                if v1 and v2 and v2 != 0:
                    lo, hi = sorted([v1, v2])
                    val = (f"{format_value(lo, {'decimals':0}, nf)} / "
                           f"{format_value(hi, {'decimals':0}, nf)}")
            cells.append(val if val not in (None, "") else "–")
        if any(c != "–" for c in cells):
            rows.append({"label": _label(schema, en, en), "unit": unit, "cells": cells})

    # tap changer
    taps = []
    tc = get_display("Tipo commutatore", raw, schema)
    if tc:
        taps.append({"label": _label(schema, "Tap changer", "Tap changer"), "value": tc})
    pp, pm = raw.get("Posizioni + rif. MT"), raw.get("Posizioni - rif. MT")
    if not is_blank(pp) or not is_blank(pm):
        taps.append({"label": _label(schema, "Tap positions (+/-)", "Tap positions (+/-)"),
                     "value": f"+{int(pp) if not is_blank(pp) else 0}/-{int(pm) if not is_blank(pm) else 0}"})
    step = raw.get("% gradino rif. MT")
    if not is_blank(step):
        try:
            v = float(step)
            v = v * 100 if v < 1 else v
            taps.append({"label": _label(schema, "Step per tap", "Step per tap"), "value": f"{format_value(v, {'decimals':2}, nf)} %"})
        except (TypeError, ValueError):
            pass

    return {"windings": labels, "rows": rows, "taps": taps}


def short_circuit_row(raw, schema):
    """Impedenza cc per Electrical, con etichette coppia:
    'MV–LV1: 6% / MV–LV2: 6% / LV1–LV2: 3,5%'."""
    nf = _nf(schema)
    windings = build_windings(raw, schema)
    rolelabel = {w["role"]: w["label"] for w in windings}
    parts = []
    if "BT2" in rolelabel:
        for key, w1, w2 in IMPEDANCES:
            if w1 in rolelabel and w2 in rolelabel:
                v = _imp(raw.get(key), nf)
                if v is not None and _f(raw.get(key)) != 0:
                    parts.append(f"{rolelabel[w1]}–{rolelabel[w2]}: {v}%")
    else:
        raw_sc = raw.get("Impedenza di cortocircuito % Totale") or raw.get("Impedenza di cortocircuito % MT-BT1")
        v = _imp(raw_sc, nf)
        if v is not None and _f(raw_sc) != 0:
            if "MT" in rolelabel and "BT1" in rolelabel:
                parts.append(f"{rolelabel['MT']}–{rolelabel['BT1']}: {v}%")
            else:
                parts.append(f"{v}%")
    if not parts:
        return None
    return {"it": None, "en": _label(schema, "Short-circuit impedance", "Short-circuit impedance"), "section": "electrical", "unit": None,
            "value": " / ".join(parts), "blank": False, "translated_ok": True}


def winding_temp_row(raw, schema, windings):
    """Sovratemperatura avvolgimenti per Working Conditions: valori separati da ' / '."""
    nf = _nf(schema)
    vals = []
    for w in windings:
        v = raw.get(w["m"]["tr"])
        if not is_blank(v):
            vals.append(format_value(v, {"decimals": 0}, nf))
    if not vals:
        return None
    return {"it": None, "en": _label(schema, "Winding temp. rise", "Winding temp. rise"), "section": "environmental", "unit": "°C",
            "value": " / ".join(vals), "blank": False, "translated_ok": True}


def standards_text(raw, family=None):
    xs = [str(raw.get(f"Norma {i} / Regol. {i}")).strip()
          for i in range(1, 5) if not is_blank(raw.get(f"Norma {i} / Regol. {i}"))]
    # IEC 60076 generico -> famiglia-specifico: -11 resina, -1 olio
    if family in ("resina", "olio"):
        suffix = "11" if family == "resina" else "1"
        xs = [f"IEC 60076-{suffix}" if x.replace(" ", "").upper() == "IEC60076" else x for x in xs]
    return " · ".join(xs) if xs else None


def earthing_rows(raw, schema):
    """Righe elettriche aggiuntive per i trasformatori di terra (earthing):
    Zo/Ro/Xo in ohm/phase + corrente di guasto omopolare al neutro."""
    if not is_earthing(raw):
        return []
    nf = _nf(schema)
    lg = schema.get("lang") or {}
    uom = (lg.get("uom") or {}).get("ohm_phase", "ohm/phase")
    def _fixed2(v):
        s = f"{_f(v):.2f}"
        return s.replace(".", ",") if nf.get("decimal") == "," else s
    rows = []
    for key in ("Zo", "Ro", "Xo"):
        v = raw.get(key)
        if not is_blank(v):
            rows.append({"it": None, "en": _label(schema, key, key), "section": "electrical",
                         "unit": uom, "value": _fixed2(v),
                         "blank": False, "translated_ok": True})
    # corrente di guasto omopolare: "300A 3s – 50A cont."
    i_fault = raw.get("Corrente guasto omopolare")
    dur = raw.get("Durata guasto omopolare")
    i_cont = raw.get("Corrente guasto permanente")
    if not is_blank(i_fault):
        cont_lbl = (lg.get("ui") or {}).get("cont", "cont.")
        parts = [f"{format_value(i_fault, {'decimals':0}, nf)}A"]
        if not is_blank(dur):
            d = _f(dur)
            parts.append(f"{format_value(d/60, {'decimals':0}, nf)}min" if d and d > 60
                         else f"{format_value(dur, {'decimals':0}, nf)}s")
        val = " ".join(parts)
        if not is_blank(i_cont):
            val += f" – {format_value(i_cont, {'decimals':0}, nf)}A {cont_lbl}"
        rows.append({"it": None,
                     "en": _label(schema, "Neutral zero-sequence fault current",
                                  "Neutral zero-sequence fault current"),
                     "section": "electrical", "unit": None, "value": val,
                     "blank": False, "translated_ok": True})
    return rows


def ip_note(raw, family, accessories, schema):
    """Nota grado IP per resina con box di protezione (estrae IPxx dagli accessori)."""
    if family != "resina":
        return None
    import re
    for a in (accessories or []):
        m = re.search(r"\bIP\s?(\d{2})\b", str(a))
        if m:
            tpl = ((schema.get("lang") or {}).get("ui") or {}).get(
                "ip_note", "Protection degree {ip} provided by the protection box.")
            return tpl.format(ip=f"IP{m.group(1)}")
    return None


def efficiency_row(raw, schema):
    val = raw.get("PEI / MEPS / HEPS in AN/ONAN")
    if is_blank(val):
        return None
    norme = " ".join(str(raw.get(f"Norma {i} / Regol. {i}") or "") for i in range(1, 5)).upper()
    eff = schema["efficiency"]
    if any(t.upper() in norme for t in eff["meps_when"]):
        label = "Efficiency index (MEPS/HEPS)"
    elif any(t.upper() in norme for t in eff["pei_when"]):
        label = "Efficiency index (PEI)"
    else:
        label = "Efficiency index"
    return {"it": None, "en": _label(schema, label, label), "section": "electrical", "unit": "%",
            "value": format_value(val, {"decimals": 3}, _nf(schema)),
            "blank": False, "translated_ok": True}


def cesi_text(raw, family, schema=None):
    if family != "resina":
        return None
    cls = raw.get("Classe amb / clim / fuoco")
    if is_blank(cls):
        return None
    suffix = "type test nr. B4013916"
    if schema and schema.get("lang"):
        suffix = schema["lang"]["ui"].get("cesi_suffix", suffix)
    return f"{str(cls).replace(' ', '')} {suffix}"


UM_THRESHOLD = 72.5  # kV, soglia prove aggiuntive


def _um_kv(raw):
    """Um in kV = primo numero della classe isolamento MT ('24 / 50 / 125' -> 24)."""
    v = raw.get("Classe isolamento MT")
    if is_blank(v):
        return None
    first = str(v).split("/")[0].strip().replace(",", ".")
    try:
        return float(first)
    except ValueError:
        return None  # es. cella corrotta in data: soglia non valutabile -> nessuna prova HV


def is_earthing(raw):
    return str(raw.get("Gruppo vettoriale") or "").strip().upper().startswith("Z")


def build_tests(raw, family, earthing_override=None):
    """Prove di routine IEC 60076-1 secondo le casistiche (dal foglio 'Principale').
    Prove interne (SFERA/DFR/insulation resistance/core demag) escluse per scelta."""
def build_tests(raw, family, earthing_override=None, schema=None):
    """Prove di routine IEC 60076-1 secondo le casistiche (dal foglio 'Principale').
    Le frasi sono prese dal glossario della lingua attiva (schema['lang']['tests']).
    Prove interne (SFERA/DFR/insulation resistance/core demag) escluse per scelta."""
    T = (schema.get("lang", {}).get("tests") if schema else None) or load_lang("en")["tests"]
    um = _um_kv(raw)
    hv = um is not None and um > UM_THRESHOLD
    oil = family == "olio"
    earth = is_earthing(raw) if earthing_override is None else earthing_override
    oltc = "sottocarico" in str(raw.get("Tipo commutatore") or "").lower()
    note_e = T["earthing_note"] if earth else ""

    t = [
        T["winding_resistance"],
        T["voltage_ratio"] + note_e,
        T["sc_impedance_load_loss"] + note_e,
        T["no_load_loss"],
        T["dielectric_header"],
        "– " + T["av"],
        "– " + (T["ivpd"] if hv else T["ivw"]),
        "– " + T["auxw"],
    ]
    if hv:
        t.append("– " + T["li"])
        t.append("– " + T["ltac"])
    if oltc:
        t.append(T["oltc"])
    if oil:
        t.append(T["tightness"])
    t.append(T["ct_check"])
    t.append(T["core_frame"])
    if not oil:                       # resina: misura scariche parziali
        t.append(T["pd_measurement"])
    if earth:
        t.append(T["zero_sequence"])
    if hv:
        t.append(T["um_header"])
        t.append("– " + T["capacitances"])
        t.append("– " + T["dc_insulation"])
        t.append("– " + T["tan_delta"])
        t.append("– " + T["dga"])
        t.append("– " + T["no_load_90_110"])
    return t


def designation(raw, schema):
    nf = _nf(schema)
    serie = str(raw.get("Serie") or "").strip()
    tail = []
    power = raw.get("Potenza nominale MT")
    if not is_blank(power):
        tail.append(f"{format_value(power, {'decimals':0}, nf)} kVA")
    hv, lv = raw.get("Tensione MT"), raw.get("Tensione BT1")
    if not is_blank(hv) and not is_blank(lv):
        tail.append(f"{format_value(hv, {'decimals':0}, nf)} / {format_value(lv, {'decimals':0}, nf)} V")
    elif not is_blank(hv):
        tail.append(f"{format_value(hv, {'decimals':0}, nf)} V")
    if tail:
        return f"{serie} — " + " · ".join(tail) if serie else " · ".join(tail)
    return serie


def designation_parts(raw, schema):
    """main = potenza; voltage = HV / LV con i valori multipli uniti da '-'
    (es. '10.000-20.000 / 720-400 V'); series = codice serie."""
    nf = _nf(schema)
    serie = str(raw.get("Serie") or "").strip()
    power = raw.get("Potenza nominale MT")
    pf = raw.get("Potenza reg. forzato MT")
    if not is_blank(power):
        p, pforced = _f(power), _f(pf)
        if pforced not in (None, 0) and pforced != p:      # doppia potenza: min-max
            lo, hi = sorted([p, pforced])
            main = (f"{format_value(lo, {'decimals':0}, nf)}-"
                    f"{format_value(hi, {'decimals':0}, nf)} kVA")
        else:
            main = f"{format_value(power, {'decimals':0}, nf)} kVA"
    else:
        main = serie or "Transformer"

    def side(keys, sort=False):
        vals = []
        for k in keys:
            v = raw.get(k)
            if not is_blank(v) and _f(v) not in (None, 0):
                vals.append(_f(v))
        if sort:
            vals = sorted(vals)
        return "-".join(format_value(v, {"decimals": 0}, nf) for v in vals)

    hv = side(["Tensione MT", "Tensione MT2"], sort=True)          # doppia tensione: dal più basso
    lv = side(["Tensione BT1", "Tensione BT2", "Tensione BT3"])    # avvolgimenti in ordine
    if hv and lv:
        voltage = f"{hv} / {lv} V"
    elif hv:
        voltage = f"{hv} V"
    else:
        voltage = ""
    return main, voltage, serie


def parse(xlsx_path, schema=None, overrides=None, lang_code="en"):
    schema = schema or load_schema()
    schema = {**schema, "lang": load_lang(lang_code)}   # glossario lingua attiva
    raw, accessories = read_raw(xlsx_path, schema)
    if overrides:
        for k, v in overrides.items():
            raw[k] = v  # corrected / supplied values from the dynamic form
    family = derive_family(raw, schema)
    ov = overrides or {}
    image_key = ov.get("__image__") or ov.get("image") or select_image(raw, family, schema)

    fields = build_fields(raw, family, schema)
    windings = build_windings(raw, schema)
    eff = efficiency_row(raw, schema)
    sci = short_circuit_row(raw, schema)
    wtr = winding_temp_row(raw, schema, windings)
    if eff:
        fields.append(eff)
    for er in earthing_rows(raw, schema):   # Zo/Ro/Xo + fault current (solo earthing)
        fields.append(er)
    if wtr:
        # inserisci winding temp rise appena PRIMA di LpA, così LpA/LwA
        # restano gli ultimi due e cadono su una riga propria
        insert_at = len(fields)
        for i, f in enumerate(fields):
            if f["it"] == "LpA":
                insert_at = i
                break
        fields.insert(insert_at, wtr)

    return {
        "raw": raw,
        "family": family,
        "image_key": image_key,
        "designation": designation(raw, schema),
        "designation_parts": designation_parts(raw, schema),
        "fields": fields,
        "sections": grouped_sections(fields, schema),
        "ratings": build_ratings(raw, schema),
        "standards": standards_text(raw, family),
        "short_circuit": sci["value"] if sci else None,
        "ip_note": ip_note(raw, family, accessories, schema),
        "cesi": cesi_text(raw, family, schema),
        "tests": build_tests(raw, family, ov.get("__earthing__"), schema),
        "accessories_excel": accessories,
        "ui": schema["lang"]["ui"],
        "sections_ui": schema["lang"]["sections"],
        "dims": {
            "L": raw.get("Lunghezza trafo"),
            "W": raw.get("Larghezza trafo"),
            "H": raw.get("Altezza trafo"),
        },
    }


def translate_accessory(text, gloss):
    """Traduce una voce accessorio mantenendo la quantità iniziale ('Nr. 3')
    e i parametri finali (sigle/numeri, es. '24kV - 250A').
    gloss = dizionario {descrizione_it_senza_quantità: traduzione}."""
    import re
    s = str(text).strip()
    if not s:
        return text
    m = re.match(r"^(Nr\.?\s*\d+\s+)(.*)$", s, flags=re.IGNORECASE)
    prefix = m.group(1) if m else ""
    body = m.group(2) if m else s
    best = None
    for k in gloss:
        if body.startswith(k) and (best is None or len(k) > len(best)):
            best = k
    if best is None:
        return text  # nessuna corrispondenza: lascia invariato
    rest = body[len(best):]  # parametri finali (kV, A, ...)
    return f"{prefix}{gloss[best]}{rest}"


def translate_color(text, color_gloss):
    """Traduce la descrizione del colore. Se il valore è solo un codice RAL
    (es. 'RAL 7030') lo lascia invariato; altrimenti sostituisce i termini noti
    mantenendo i codici RAL e i numeri."""
    import re
    s = str(text).strip()
    if not s:
        return text
    if re.match(r"^RAL\s*\d+\w*$", s, flags=re.IGNORECASE):
        return text  # solo RAL: non tradurre
    out = s
    for term in sorted(color_gloss, key=len, reverse=True):
        out = re.sub(re.escape(term), color_gloss[term], out, flags=re.IGNORECASE)
    return out
